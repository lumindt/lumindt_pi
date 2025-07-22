import time
from datetime import datetime
from threading import Thread
from adafruit_ads1x15.ads1115 import ADS1115
from adafruit_ads1x15.analog_in import AnalogIn
import board
import busio
from openpyxl import Workbook, load_workbook
import os
import math

# === CONFIGURATION ===
V_SUPPLY = 2        # Supplied voltage to bridge (V)
R_FIXED = 10      # Fixed resistor (Ohms)
LOG_INTERVAL = 0.01  # Time between samples (seconds)
EXCEL_FILE = "resistance_data2.xlsx"

# === GLOBAL STOP FLAG ===
stop_flag = False

def input_listener():
    global stop_flag
    while True:
        user_input = input()
        if user_input.strip().lower() == 'stop':
            stop_flag = True
            break

# === ADC SETUP ===
i2c = busio.I2C(board.SCL, board.SDA)
adc = ADS1115(i2c)
adc.gain = 1
channel = AnalogIn(adc, 0, 1)  # Differential mode between A0 and A1

# === EXCEL SETUP ===
if os.path.exists(EXCEL_FILE):
    wb = load_workbook(EXCEL_FILE)
    existing_sheets = wb.sheetnames
    run_numbers = [
        int(name.replace("Run_", "")) for name in existing_sheets if name.startswith("Run_")
    ]
    next_run = max(run_numbers, default=0) + 1
else:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    next_run = 1

sheet_name = f"Run_{next_run}"
ws = wb.create_sheet(title=sheet_name)
ws.append(["Elapsed ms", "time_sec","Voltage (V)", "R_var"])

# === START INPUT LISTENER THREAD ===
listener_thread = Thread(target=input_listener, daemon=True)
listener_thread.start()

# === DATA LOGGING LOOP ===
print(f"\nStarted logging to sheet '{sheet_name}' — type 'stop' and press ENTER to end.\n")
# === CAPTURE START TIME ===
start_time = time.time()
try:
    while not stop_flag:          
        v_out = channel.voltage
        try:
            x = (R_FIXED / (R_FIXED + R_FIXED)) + (v_out / V_SUPPLY)
            r_var = (R_FIXED * x) / (1 - x)
            
        except ZeroDivisionError:
            r_var = float('inf')

        elapsed_ms = (time.time() - start_time) * 1000  # ms since start
        time_sec = elapsed_ms / 1000
        

        print(f"[{elapsed_ms:.0f} ms] |time_s:{time_sec if time_sec is not None else 'N/A'} | Voltage: {v_out:.4f} V | R_var: {r_var:.6f} Ω")

        ws.append([round(elapsed_ms, 2), time_sec, round(v_out, 4), round(r_var, 6)])
        time.sleep(LOG_INTERVAL)

finally:
    wb.save(EXCEL_FILE)
    print(f"\nLogging stopped. Data saved to '{EXCEL_FILE}' in sheet '{sheet_name}'.")
